from openpyxl import load_workbook
from datetime import date
from XMLToExcel_NFe import ToExcel
from search_status_nfe import DB_search

def SearchDB(nFE_list):
    return DB_search(nFE_list)

def OverWriteExcel():
    print("📥 Lendo XMLs e convertendo para DataFrame...")
    df = ToExcel()

    print("🔍 Consultando status no banco de dados...")
    df = SearchDB(df)

    # (Opcional) Reorganiza colunas em ordem desejada, se existirem
    desired_order = ['Id', 'Nome', 'Emissao', 'NF', 'Serie', 'ChaveNF', 'CFOP', 'STATUS']
    df = df[[col for col in desired_order if col in df.columns]]

    datetoday = date.today().strftime("%d-%b-%Y")

    try:
        print("📂 Abrindo planilha existente...")
        workbook = load_workbook("NFe_distribuidoras.xlsx")

        # Remove a aba do dia, se já existir
        if datetoday in workbook.sheetnames:
            print(f"⚠️ Aba '{datetoday}' já existe. Substituindo...")
            del workbook[datetoday]

        # Cria nova aba com a data de hoje
        new_sheet = workbook.create_sheet(datetoday)

        print("✍️ Escrevendo dados na nova aba...")
        new_sheet.append(df.columns.tolist())  # cabeçalho

        for row in df.itertuples(index=False):
            new_sheet.append(list(row))

        workbook.save("NFe_distribuidoras.xlsx")
        print("✅ Planilha atualizada com sucesso!")

    except PermissionError:
        print("❌ Erro: Feche o arquivo 'NFe_distribuidoras.xlsx' antes de executar este script.")
    except FileNotFoundError:
        print("❌ Arquivo 'NFe_distribuidoras.xlsx' não encontrado.")
    except Exception as e:
        print(f"❌ Ocorreu um erro inesperado: {e}")